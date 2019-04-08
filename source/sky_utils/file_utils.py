import logging
import openpyxl
import pandas as pd
import time
import xlsxwriter
from pdb import set_trace
from openpyxl.styles import Alignment, Font, Color
from openpyxl.styles.borders import Border, Side
from openpyxl.drawing.image import Image
from sky_utils import email_module as em

TIMESTAMP = time.strftime("%Y%m%d_%H%M%S")
ICON_DIR = "D:\\report_manager\\Icons\\"

def generate_selections_dataframe(sel_file_path):
    df_sel = pd.read_excel (sel_file_path, header=None)
    selections = []
    s = []
    selections = df_sel.values.tolist()
    selections = [x for x in selections[0] if str(x) != 'nan']
    s = selections[0].split('\n')
    s = list(filter(None, s))
    s_len = len(s)
    df_sel = pd.DataFrame({'selections': s})
    df_sel = pd.DataFrame(df_sel.selections.str.split(':',1).tolist(), columns = ['labels','selections'])

    return df_sel, s_len, sel_file_path


def email_reports(client_id, filename, report_type, email_id):
    from_email = 'support'
    subject = client_id + ' ' + 'StyleSelling' + ' ' + 'Report'
    attachment = filename
    message = 'Please' + ' ' + 'See' + ' ' + 'Attached.'

    if report_type == 'OnDemandExport':
        recipients = email_id
        logging.debug("Sending email")
        em.send_email_from(from_email, [recipients], subject, message, attachment)
    if report_type == 'ScheduledExport':
        recipients = ["sbellala@skyitgroup.com"]
        logging.debug("Sending email")
        em.send_email_from(from_email, recipients, subject, message, attachment)


def format_info(wb, s_len, client_id):

    logger = logging.getLogger('openpyxl.drawing.image')
    logging.getLogger("Image").setLevel(logging.WARNING)

    icon_path = ICON_DIR + client_id + "_Icon.png"
    wb.active = 1
    ws = wb["Info"]

    # updating dimensions for rows and columns
    ws.row_dimensions[1].height = 52
    ws.row_dimensions[2].height = 30
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18

    # Merging 1st and 2nd row
    ws.merge_cells(start_row = 1, start_column = 1, end_row = 1, end_column = 4)
    ws.merge_cells(start_row = 2, start_column = 1, end_row = 2, end_column = 4)

    # formatting style selling report cell
    ss_font = Font(size = 20) #ss --> style selling
    ss_cell = ws.cell(2,1)
    ss_cell.value = 'Style Selling Report'
    ss_cell.font = ss_font
    ss_cell.alignment = Alignment(horizontal='center', vertical='center')

    # formatting Report Generated: cell
    rg_font = Font(bold=True, size = 11) #rg --> Report Generated
    rg_cell = ws.cell(3,1)
    rg_cell.value = 'Report Generated:'
    rg_cell.font = rg_font

    # adding Timestamp   20190314_133248
    date_time = TIMESTAMP[4:6] + '/' + TIMESTAMP[6:8] + '/' + TIMESTAMP[0:4] + ' ' + TIMESTAMP[9:11] + ':' + TIMESTAMP[11:13]
    ws.cell(3,2).value = date_time

    # formatting Selections: cell
    se_font = Font(bold=True, size = 11) #se --> Selections:
    se_cell = ws.cell(5,1)
    se_cell.value = 'Selections:'
    se_cell.font = se_font

    # formatting Selections labels
    se_labels_font = Font(bold=True, size = 11)
    for cell in range(5, 5+s_len):
        cell = ws.cell(cell, 2)
        cell.font = se_labels_font

    # inserting Client icon
    cell_coord = ws.cell(row=1, column=2).coordinate
    img = openpyxl.drawing.image.Image(icon_path)
    ws.add_image(img, cell_coord)

    return wb
