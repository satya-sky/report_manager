
import image
import io
import logging
import openpyxl
import os
import pandas as pd
import pdb
import shutil
import sys
import time
import urllib3
import xlsxwriter
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, Color
from openpyxl.styles.borders import Border, Side
from sky_utils import email_module as em
from sky_utils import file_utils as futil


TIMESTAMP = time.strftime("%Y%m%d_%H%M%S")

# production directories
FILE_DIR = ('\\\\192.168.100.25\\Qlik\\ReportManager_Test\\')
ROOT_DIR = 'D:\\report_manager\\'
OUTPUT_DIR = "D:\\report_manager\\output\\"

# local directories
# FILE_DIR = "C:\\report_manager\\001files\\"
# ROOT_DIR = "C:\\report_manager\\"
# OUTPUT_DIR = "C:\\report_manager\\output\\"

def generate_cls_report(file_path, filename, client_id, report_type, file_id, sel_filename, email_id):

    # pdb.set_trace()
    if file_path.endswith(".xls") or file_path.endswith(".xlsx"):
       df = pd.read_excel(file_path)    #creating dataframe for data sheet
       os.remove(file_path)

       # findng info sheet for available data sheet
       for filename in os.listdir(FILE_DIR + "Selections\\"):
           sel_file_path = FILE_DIR + "Selections\\" + filename
           if filename == sel_filename:
               logging.debug(TIMESTAMP + "||" + file_id + "||" + "Selections file found " + sel_filename)
               # creating dataframe for Info sheet
               df_sel, s_len, sel_file_path = futil.generate_selections_dataframe(sel_file_path)
               os.remove(sel_file_path)

       # Creating output file and writing dataframes for Info sheet and data sheet
       output_file = client_id + '_Output_' + file_id +'.xlsx'
       writer = pd.ExcelWriter(output_file,engine='xlsxwriter')
       df_sel.to_excel(writer, sheet_name='Info', startrow=4, startcol=1, header=False, index=False)
       df.to_excel(writer, sheet_name='StyleSelling', index=False)
       rows = df.shape[0] + 1
       columns = df.shape[1] + 1
       workbook = writer.book
       output_file = workbook.filename
       writer.save()

       # opening workbook in openpyxl to start formatting
       wb = openpyxl.load_workbook(output_file)

       # formatting Info Sheet
       logging.debug(TIMESTAMP + "||" + file_id + "||" + ": Info Sheet formatting started")
       wb = futil.format_info(wb, s_len, client_id)
       logging.debug(TIMESTAMP + "||" + file_id + "||" + ": Info Sheet formatting completed")

       # formatting Data Sheet rows and columns with required dimensions
       logging.debug("Data Sheet formatting started")
       wb.active = 2
       ws = wb["StyleSelling"]
       ws.sheet_view.showGridLines = False
       ws.row_dimensions[1].height = 30
       ws.column_dimensions['A'].width = 20.75
       ws.column_dimensions['N'].width = 20.75

       for col in range(2, columns-1):
           cell_coord = ws.cell(row = 1, column = col).coordinate[0]
           ws.column_dimensions[cell_coord].width = 10.75

       # creating different border variables needed for formatting various cells
       thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

       thick_border = Border(left=Side(style='thick'),
                      right=Side(style='thick'),
                      top=Side(style='thick'),
                      bottom=Side(style='thick'))

       ver_thick_border = Border(left=Side(style='thick'),
                          right=Side(style='thick'),
                          top=Side(style=None),
                          bottom=Side(style=None))

       hor_thick_border = Border(left=Side(style=None),
                          right=Side(style=None),
                          top=Side(style='thick'),
                          bottom=Side(style='thick'))

       brc_thick_border = Border(left=Side(style=None),
                          right=Side(style='thick'),
                          top=Side(style=None),
                          bottom=Side(style='thick'))

       blc_thick_border = Border(left=Side(style=None),
                          right=Side(style='thick'),
                          top=Side(style=None),
                          bottom=Side(style='thick'))

       # Formatting Column headers
       header_clr = openpyxl.styles.colors.Color(rgb='00336699')
       header_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=header_clr)
       header_font = Font(color='00FFFFFF')
       for col in range(1,columns):
            cell = ws.cell(row=1, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText= True)
            cell.fill = header_fill
            cell.font = header_font

       # Merging 'Pattern Name' Column and aligning
       pattern_clr = openpyxl.styles.colors.Color(rgb='00B5E2FF')
       pattern_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=pattern_clr)
       pattern_font = Font(size = 13)
       for i in range(2,rows):
           for j in range(i,rows):
                if ws.cell(row = j, column = 1).value == ws.cell(row = j+1, column = 1).value:
                    j = j+1
                else:
                    ws.merge_cells(start_row = i, start_column = 1, end_row = j, end_column = 1)
                    cell = ws.cell(row=i, column=1)
                    cell.border = thick_border
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    cell.fill = pattern_fill
                    cell.font = pattern_font
                    i = j
                    break

       # Merging 'Style' Column and aligning
       style_font = Font(size = 8)
       for i in range(2,rows):
           for j in range(i,rows):
                if ws.cell(row = j, column = 2).value == ws.cell(row = j+1, column = 2).value:
                    j = j+1
                else:
                    ws.merge_cells(start_row = i, start_column = 2, end_row = j, end_column = 2)
                    cell = ws.cell(row=i, column=2)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = style_font
                    i = j
                    break

       # formatting color column
       color_font = Font(size = 14)
       for row in range(2,rows):
           cell = ws.cell(row = row, column = 3)
           cell.font = color_font

       # formatting Material_desc column
       material_font = Font(size = 8)
       for row in range(2,rows):
           cell = ws.cell(row = row, column = 4)
           cell.font = material_font

       # formatting numbers
       numbers_font = Font(size = 14)
       for col in range(5,columns-1):
           for row in range(2,rows+3):
               cell = ws.cell(row = row, column = col)
               cell.font = numbers_font

       # Merging 'Image' Column and aligning
       for i in range(2,rows+1):
           for j in range(i,rows+1):
                if ws.cell(row = j, column = 14).value == ws.cell(row = j+1, column = 14).value:
                    j = j+1
                else:
                    ws.merge_cells(start_row = i, start_column= 14, end_row = j, end_column = 14)
                    cell = ws.cell(row=i, column=14)
                    cell.border = ver_thick_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    i = j
                    break

       # Borders for remaining Image columns
       for row in range(2,rows):
           cell = ws.cell(row = row, column = 14)
           cell.border = ver_thick_border

       # inserting Image loop
       logging.getLogger("urllib3").setLevel(logging.WARNING)
       logger = logging.getLogger('Image')
       for row in range(2, rows):
           cell = ws.cell(row = row, column = 14).value
           cell_coord = ws.cell(row=row, column=14).coordinate
           if cell == None:
               row = row + 1
           else:
               http = urllib3.PoolManager()
               r = http.request('GET', cell)
               image_file = io.BytesIO(r.data)
               img = Image(image_file)
               ws.add_image(img, cell_coord)
               # ws.cell.aligning = WD_ALIGN_PARAGRAPH.CENTER
               ws.cell(row = row, column = 14).value = None

       # updating row height and adding borders to Totals
       # Cells(j, 14).Interior.Color = RGB(255, 255, 255)
       total_clr = openpyxl.styles.colors.Color(rgb='00C4C2C0')
       total_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=total_clr)
       total_font = Font(bold=True, size = 14)
       for row in range(2,rows):
            if ws.cell(row = row, column = 2).value == 'Total':
                ws.row_dimensions[row].height = 20
                for col in range(2, columns-1):
                    cell = ws.cell(row=row, column=col)
                    cell.border = hor_thick_border
                    cell.fill = total_fill
                    cell.font = total_font
            else:
                ws.row_dimensions[row].height = 35

       # updating font of Style total cells
       style_total_font = Font(bold=True, size = 12)
       for row in range(2,rows):
            if ws.cell(row = row, column = 2).value == 'Total':
                cell = ws.cell(row=row, column=2)
                cell.font = style_total_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

       # updating main total row height and adding border
       main_total_font = Font(bold=True, size = 14)
       for row in range(2,rows+1):
            if ws.cell(row = row, column = 1).value == 'Total':
                ws.row_dimensions[row].height = 35
                for col in range(1, columns-1):
                    cell = ws.cell(row=row, column=col)
                    cell.border = hor_thick_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = main_total_font
            else:
                pass

       # formatting Main total row
       main_tot_font = Font(size = 12, bold = True)
       if ws.cell(row = rows, column = 1).value == 'Total':
           cell = ws.cell(row = rows, column = 1)
           cell.font = main_tot_font
           cell.alignment = Alignment(horizontal='left', vertical='center')
           cell.border = ver_thick_border

       # formatting Main total cell (Image column)
       # brc = botton right corner
       brc_cell = ws.cell(row = rows, column = columns-1)
       brc_cell.border = brc_thick_border
       brc_cell.value = None

       # formatting Main total cell (column 1)
       # blc = botton left corner
       blc_cell = ws.cell(row = rows, column = 1)
       blc_cell.border = thick_border

       # Aligning cells to center
       for column in range(3, columns-1):
           for row in range(2, rows):
               cell = ws.cell(row=row, column=column)
               cell.alignment = Alignment(horizontal='center', vertical='center')

       # format LW ST% column)
       for row in range(2,rows+1):
           for col in range(columns-3,columns):
               cell = ws.cell(row = row, column = col)
               cell.number_format = '0.0%'

       # format LW AUR column (currency))
       for row in range(2,rows+1):
           cell = ws.cell(row = row, column = 6)
           # cell.number_format = '"$"#,##0.0_-'
           cell.number_format = '"$"#,##0.0_);("$"#,##0.0)'

       # format LW Sales $ (currency))
       for row in range(2,rows+1):
           cell = ws.cell(row = row, column = 7)
           # cell.number_format = '"$"#,##0_-'
           cell.number_format = '"$"#,##0_);("$"#,##0)'

       # format STD AUR
       for row in range(2,rows+1):
           cell = ws.cell(row = row, column = 9)
           # cell.number_format = '"$"#,##0.0_-'
           cell.number_format = '"$"#,##0.0_);("$"#,##0.0)'

       # format STD Sales $
       for row in range(2,rows+1):
           cell = ws.cell(row = row, column = 10)
           # cell.number_format = '"$"#,##0_-'
           cell.number_format = '"$"#,##0_);("$"#,##0)'

    wb.save(output_file)
    logging.debug(TIMESTAMP + "||" + file_id + "||" + ": Data Sheet formatting completed")

    # creating folder and copying output file from root folder
    # pdb.set_trace()
    output_path = OUTPUT_DIR + client_id + "\\" + file_id
    os.mkdir(output_path)
    # output_file_path = ROOT_DIR + "\\"  + output_file
    output_file_path = ROOT_DIR + "source\\" + output_file
    final_output = output_path + "\\CLSStyleSelling.xlsx"
    shutil.move(output_file_path, final_output)

    # sending out email
    futil.email_reports(client_id, final_output, report_type, email_id)
    # pdb.set_trace()
    logging.debug(TIMESTAMP + "||" + file_id + "||" + ": Email sent to " + email_id + " with attached report")

    # remove source files, output files and output folder
    # os.remove(output_file_path)
    # os.remove(sel_file_path)
    # os.remove(file_path)
    shutil.rmtree(output_path)
    logging.debug(TIMESTAMP + "||" + file_id + "||" + ": Deleted source files, temp files and temp folders")


if __name__ == "__main__":

    file_path = sys.argv[1]
    # file_path = "\\\\192.168.100.25\\Qlik\\ReportManager_Test\\CLS_OnDemandExport_sbellala@skyitgroup_01-18-2019_275.xls"
    split_file_path = file_path.split('\\')

    # extracting necessary information from filename
    filename = split_file_path[-1]
    email_id = filename.split("_")[2] + ".com"
    client_id = filename.split('_')[0]
    report_type = filename.split('_')[1]
    file_id = filename.split('_')[-1].replace('.xls','')
    sel_filename = filename.replace('OnDemandExport','Selections')

    # creating logging file
    logging.basicConfig(filename= ROOT_DIR + 'log\\' + client_id + '_ReportManager_log.txt',level=logging.DEBUG)
    logging.debug('Python file called' + time.strftime("%Y%m%d_%H%M%S"))

    # creating reports
    if client_id == 'CLS':
        generate_cls_report(file_path, filename, client_id, report_type, file_id, sel_filename, email_id)

    elif client_id == 'NIZ':
        generate_niz_report(file_path, filename, client_id, report_type, file_id, sel_filename, email_id)

    else:
        logging.debug("No Files Present")
